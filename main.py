from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from tkinter import *
from tkinter import messagebox
import win32com.client
from threading import Thread
import time
from datetime import datetime
import re
import openpyxl

def carregar_cpfs(caminho_arquivo):
    workbook = openpyxl.load_workbook(caminho_arquivo)
    sheet = workbook.active
    cpfs = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Supõe que os CPFs estão na primeira coluna começando da segunda linha
        cpfs.append(row[0])
    return cpfs

def fechar_aviso(navegador):
    try:
        wait = WebDriverWait(navegador, 5)
        botao_aviso = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/div[2]/div[1]/button')))
        botao_aviso.click()
    except NoSuchElementException:
        pass  # Se o aviso não estiver presente, apenas continua o código

def fazer_login(cpfs):
    options = Options()
    options.headless = False
    navegador = webdriver.Firefox(options=options)
    navegador.get('https://#')
    wait = WebDriverWait(navegador, 20)
    matricula = matricula_entry.get()
    senha = senha_entry.get()

    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="username"]'))).send_keys(matricula)
    time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="password"]'))).send_keys(senha)
    time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="kc-login"]'))).click()
    time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[2]/aside/section/ul/li[5]/a/span'))).click()

    for i, cpf in enumerate(cpfs, start=2):  # Começa da linha 2, onde os CPFs estão localizados
        # Pesquisa por CPF:
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[3]/section/div/div/div[2]/div/form/div/div[1]/input'))).clear()
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[3]/section/div/div/div[2]/div/form/div/div[1]/input'))).send_keys(cpf)
        time.sleep(2)
        wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[3]/section/div/div/div[2]/div/form/div/div[2]/button'))).click()
        time.sleep(2)

        # Salvar as informações antes da verificação da frase para opt-in
        salvar_informacoes(navegador, i)

        # Verificação da presença da frase para opt-in
        if "Cliente não recebeu solicitação para opt-in" in navegador.page_source:
            # Continuar se a frase estiver presente:
            wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div[3]/section/div/div/div[2]/div/div[3]/div/div[2]/div/div[2]/div[2]/div/table/tbody/tr/td[5]/input'))).click()
            time.sleep(2)
            wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[3]/section/div/div/div[2]/div/div[3]/div/div[2]/div/div[1]/div/button'))).click()
            salvar_status(navegador, i, True)  # Salva 'SIM' na coluna H
            time.sleep(2)
        else:
            salvar_status(navegador, i, False)  # Salva 'NÃO' na coluna H se a frase não for encontrada
            continue  # Passa para o próximo CPF se a frase não for encontrada

        # Segunda verificação de frase
        time.sleep(2)
        if "Cliente/Conta sem adesão ao SIGMS. Por favor, orientar a realização da adesão para prosseguir com o atendimento." in navegador.page_source:
            wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div/div[1]/button/span'))).click()
            salvar_status(navegador, i, False)  # Salva 'NÃO' na coluna H se a frase não for encontrada
            time.sleep(2)
            continue # Passa para o próximo CPF se a frase não for encontrada
        else:
            # Encontrar todas as linhas com os dados de telefone
            time.sleep(2)
            linhas = navegador.find_elements(By.CSS_SELECTOR, "tr[ng-repeat='item in vm.listaTelefones']")

            # Preparar para extrair a data mais recente e o elemento correspondente
            data_mais_recente = None
           elemento_mais_recente = None

            for linha in linhas:
                # Extrair o texto da linha e encontrar a data
                texto = linha.find_element(By.CSS_SELECTOR, "td").text
                data_str = texto.split('- Data Adesão: ')[1]
                data_obj = datetime.strptime(data_str, '%d/%m/%Y')

                # Comparar datas para encontrar a mais recente
                if data_mais_recente is None or data_obj > data_mais_recente:
                    data_mais_recente = data_obj
                    elemento_mais_recente = linha.find_element(By.CSS_SELECTOR, 'input[type="radio"]')

            # Clicar no elemento mais recente
            elemento_mais_recente.click()
            time.sleep(2)
            wait.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div[1]/div/div/div[3]/span/button'))).click()
            fechar_aviso(navegador)  # Fecha o aviso se estiver presente
            salvar_status(navegador, i, True)  # Salva 'SIM' na coluna H
            time.sleep(1)

            # Retorna ao início para pesquisar o próximo CPF
            continue

def salvar_informacoes(navegador, linha):
    # Tenta encontrar o primeiro elemento
    time.sleep(2)
    try:
        informacoes_element_1 = navegador.find_element(By.XPATH, '//*[@id="app"]/div/div[3]/section/div/div/div[2]/div/div[2]/div[2]/p[2]')
        informacoes_texto_1 = informacoes_element_1.text
    except NoSuchElementException:
        informacoes_texto_1 = 'NÃO HÁ'

    # Tenta encontrar o segundo elemento
    time.sleep(2)
    try:
        informacoes_element_2 = navegador.find_element(By.XPATH, '//*[@id="app"]/div/div[3]/section/div/div/div[2]/div/div[2]/div[2]/p[3]')
        informacoes_texto_2 = informacoes_element_2.text
    except NoSuchElementException:
        informacoes_texto_2 = 'NÃO HÁ'

    # Salvar as informações no arquivo Excel
    time.sleep(2)
    workbook = openpyxl.load_workbook("Base Teste.xlsx")
    sheet = workbook.active

    # Escrever as informações nas colunas J e K da linha correspondente ao CPF
    time.sleep(2)
    sheet.cell(row=linha, column=10, value=informacoes_texto_1)  # Coluna J
    sheet.cell(row=linha, column=11, value=informacoes_texto_2)  # Coluna K

    # Salvar as alterações no arquivo Excel
   workbook.save("Base Teste.xlsx")

def salvar_status(navegador, linha, status):
    workbook = openpyxl.load_workbook("Base Teste.xlsx")
    sheet = workbook.active

    # Escrever 'SIM' ou 'NÃO' na coluna H da linha correspondente ao CPF
    if status:
        sheet.cell(row=linha, column=8, value='SIM')  # Coluna H
    else:
        sheet.cell(row=linha, column=8, value='NÃO')  # Coluna H

    # Salvar as alterações no arquivo Excel
    workbook.save("Base Teste.xlsx")

# Interface gráfica
janela = Tk()
janela.title("Análise de Contratos Habitacionais - Beta")
janela.geometry("500x300")

texto_email = Label(janela, text="Matrícula:")
texto_email.grid(column=0, row=1)
matricula_entry = Entry(janela, width=25)
matricula_entry.grid(column=1, row=1)

texto_senha = Label(janela, text="Senha:")
texto_senha.grid(column=0, row=2)
senha_entry = Entry(janela, width=25, show='*')
senha_entry.grid(column=1, row=2)

botao = Button(janela, text="Pesquisar", command=lambda: Thread(target=lambda: fazer_login(carregar_cpfs("Base Teste.xlsx"))).start())
botao.grid(column=1, row=4)

janela.mainloop()


